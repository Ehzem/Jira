# # $env:JIRA_BASE_URL="https://your-site.atlassian.net" 
# # $env:JIRA_EMAIL="your-email@example.com"
# # $env:JIRA_API_TOKEN="your-api-token"

# # python ".\jira_component_epic_task_import.py" "sepm-jira.xlsx" --project-key PER


# #!/usr/bin/env python3
# """Import Jira Components, Epics, and Tasks from a three-column Excel sheet.

# Expected columns (case/spacing insensitive): components | epics | tasks

# The script is a preview unless --apply is supplied. Existing Jira work items are
# matched by exact summary inside the selected project and updated, making reruns
# safe. Jira credentials are read from environment variables:

#     JIRA_BASE_URL, JIRA_EMAIL, JIRA_API_TOKEN
# """

# from __future__ import annotations

# import argparse
# import base64
# import json
# import os
# import re
# import sys
# import time
# from collections import OrderedDict
# from dataclasses import dataclass
# from pathlib import Path
# from typing import Any, Iterable
# from urllib.error import HTTPError, URLError
# from urllib.parse import urlencode
# from urllib.request import Request, urlopen

# from openpyxl import load_workbook


# def clean(value: Any) -> str:
#     if value is None:
#         return ""
#     return re.sub(r"\s+", " ", str(value)).strip()


# def normalized_header(value: Any) -> str:
#     return re.sub(r"[^a-z0-9]", "", clean(value).casefold())


# def unique(values: Iterable[str]) -> list[str]:
#     return list(OrderedDict.fromkeys(values))


# @dataclass(frozen=True)
# class Row:
#     excel_row: int
#     component: str
#     epic: str
#     task: str


# def prompt_for_value(row_number: int, field: str, context: str) -> str:
#     """Ask for a required relationship without changing the workbook."""
#     try:
#         import tkinter as tk
#         from tkinter import simpledialog

#         root = tk.Tk()
#         root.withdraw()
#         root.attributes("-topmost", True)
#         try:
#             while True:
#                 value = simpledialog.askstring(
#                     "Missing Jira mapping",
#                     f"Excel row {row_number}: {context}\n\nEnter the missing {field}:",
#                     parent=root,
#                 )
#                 if value is None:
#                     raise ValueError(
#                         f"Import cancelled while asking for {field} in Excel row {row_number}."
#                     )
#                 value = clean(value)
#                 if value:
#                     return value
#         finally:
#             root.destroy()
#     except ValueError:
#         raise
#     except Exception as exc:
#         raise ValueError(
#             f"Excel row {row_number} requires a {field}, but the input box could not open: {exc}"
#         ) from exc


# def read_rows(path: Path, sheet_name: str | None) -> tuple[str, list[Row]]:
#     workbook = load_workbook(path, read_only=True, data_only=True)
#     try:
#         if sheet_name:
#             if sheet_name not in workbook.sheetnames:
#                 raise ValueError(
#                     f"Sheet {sheet_name!r} was not found. Available sheets: "
#                     + ", ".join(workbook.sheetnames)
#                 )
#             sheet = workbook[sheet_name]
#         else:
#             sheet = workbook[workbook.sheetnames[0]]

#         header_row = None
#         columns: dict[str, int] = {}
#         aliases = {
#             "components": "component",
#             "component": "component",
#             "epics": "epic",
#             "epic": "epic",
#             "tasks": "task",
#             "task": "task",
#         }
#         for row_number, cells in enumerate(
#             sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 30), values_only=True),
#             start=1,
#         ):
#             found: dict[str, int] = {}
#             for column_number, value in enumerate(cells, start=1):
#                 canonical = aliases.get(normalized_header(value))
#                 if canonical:
#                     found[canonical] = column_number
#             if {"component", "epic", "task"}.issubset(found):
#                 header_row = row_number
#                 columns = found
#                 break

#         if header_row is None:
#             raise ValueError(
#                 "Could not find Component, Epic, and Task columns in the first 30 rows."
#             )

#         rows: list[Row] = []
#         seen_task_mappings: set[tuple[str, str, str]] = set()
#         for row_number in range(header_row + 1, sheet.max_row + 1):
#             component = clean(sheet.cell(row_number, columns["component"]).value)
#             epic = clean(sheet.cell(row_number, columns["epic"]).value)
#             task = clean(sheet.cell(row_number, columns["task"]).value)
#             if not component and not epic and not task:
#                 continue

#             # Component-only rows are valid. Epic rows may omit Task, but every
#             # Epic needs a Component. A Task needs both its parent Epic and its
#             # Component. Prompt only for relationships that are actually required.
#             if epic and not component:
#                 component = prompt_for_value(
#                     row_number, "Component", f"Epic {epic!r} has no Component."
#                 )
#             if task and not epic:
#                 epic = prompt_for_value(
#                     row_number, "Epic", f"Task {task!r} has no parent Epic."
#                 )
#             if task and not component:
#                 component = prompt_for_value(
#                     row_number, "Component", f"Task {task!r} has no Component."
#                 )

#             item = Row(row_number, component, epic, task)
#             mapping = (task.casefold(), epic.casefold(), component.casefold())
#             if task and mapping in seen_task_mappings:
#                 continue
#             if task:
#                 seen_task_mappings.add(mapping)
#             rows.append(item)

#         if not rows:
#             raise ValueError("No data rows were found below the headers.")
#         return sheet.title, rows
#     finally:
#         workbook.close()


# class Jira:
#     def __init__(self, base_url: str, email: str, token: str) -> None:
#         self.base_url = base_url.rstrip("/")
#         credentials = base64.b64encode(f"{email}:{token}".encode("utf-8")).decode("ascii")
#         self.headers = {
#             "Authorization": f"Basic {credentials}",
#             "Accept": "application/json",
#             "Content-Type": "application/json",
#         }

#     def request(self, method: str, path: str, **kwargs: Any) -> Any:
#         params = kwargs.pop("params", None)
#         payload = kwargs.pop("json", None)
#         if kwargs:
#             raise TypeError(f"Unsupported request options: {', '.join(kwargs)}")
#         url = self.base_url + path
#         if params:
#             url += "?" + urlencode(params)
#         body = json.dumps(payload).encode("utf-8") if payload is not None else None
#         for attempt in range(5):
#             try:
#                 request = Request(url, data=body, headers=self.headers, method=method)
#                 with urlopen(request, timeout=120) as response:
#                     content = response.read()
#                     if response.status == 204 or not content:
#                         return None
#                     return json.loads(content.decode("utf-8"))
#             except HTTPError as exc:
#                 content = exc.read()
#                 try:
#                     detail = json.loads(content.decode("utf-8"))
#                 except (ValueError, UnicodeDecodeError):
#                     detail = content.decode("utf-8", errors="replace")
#                 if exc.code == 429 or 500 <= exc.code < 600:
#                     if attempt < 4:
#                         retry_after = exc.headers.get("Retry-After")
#                         time.sleep(float(retry_after) if retry_after else 2 ** attempt)
#                         continue
#                 raise RuntimeError(
#                     f"Jira {method} {path} returned HTTP {exc.code}: {detail}"
#                 ) from exc
#             except (URLError, TimeoutError) as exc:
#                 if attempt == 4:
#                     raise RuntimeError(f"Jira request failed after retries: {exc}") from exc
#                 time.sleep(2 ** attempt)
#         raise AssertionError("unreachable")

#     def project(self, key: str) -> dict[str, Any]:
#         return self.request("GET", f"/rest/api/3/project/{key}")

#     def components(self, project_key: str) -> list[dict[str, Any]]:
#         return self.request("GET", f"/rest/api/3/project/{project_key}/components")

#     def create_component(self, project_key: str, name: str) -> dict[str, Any]:
#         return self.request(
#             "POST", "/rest/api/3/component", json={"name": name, "project": project_key}
#         )

#     def issues(self, project_key: str, issue_type: str) -> list[dict[str, Any]]:
#         # Enhanced search is paginated with nextPageToken.
#         jql_type = issue_type.replace('"', '\\"')
#         jql = f'project = "{project_key}" AND issuetype = "{jql_type}" ORDER BY key'
#         result: list[dict[str, Any]] = []
#         token: str | None = None
#         while True:
#             params: dict[str, Any] = {
#                 "jql": jql,
#                 "fields": "summary,components,parent",
#                 "maxResults": 100,
#             }
#             if token:
#                 params["nextPageToken"] = token
#             page = self.request("GET", "/rest/api/3/search/jql", params=params)
#             result.extend(page.get("issues", []))
#             token = page.get("nextPageToken")
#             if not token:
#                 return result

#     def create_issue(self, fields: dict[str, Any]) -> dict[str, Any]:
#         return self.request("POST", "/rest/api/3/issue", json={"fields": fields})

#     def update_issue(self, key: str, fields: dict[str, Any]) -> None:
#         self.request("PUT", f"/rest/api/3/issue/{key}", json={"fields": fields})


# def exact_summary_index(issues: list[dict[str, Any]], issue_type: str) -> dict[str, dict[str, Any]]:
#     index: dict[str, dict[str, Any]] = {}
#     for issue in issues:
#         summary = clean(issue.get("fields", {}).get("summary"))
#         identity = summary.casefold()
#         if identity in index:
#             raise RuntimeError(
#                 f"Multiple Jira {issue_type} items have the exact summary {summary!r}. "
#                 "Rename or remove the duplicate before importing."
#             )
#         index[identity] = issue
#     return index


# def component_ids(items: Iterable[str], component_by_name: dict[str, dict[str, Any]]) -> list[dict[str, str]]:
#     return [{"id": str(component_by_name[name.casefold()]["id"])} for name in items]


# def main() -> int:
#     parser = argparse.ArgumentParser(
#         description="Import Components, Epics, and Tasks from Excel into Jira Cloud."
#     )
#     parser.add_argument("workbook", type=Path)
#     parser.add_argument("--project-key", required=True)
#     parser.add_argument("--sheet", help="Worksheet name; defaults to the first sheet")
#     parser.add_argument("--epic-type", default="Epic", help="Jira Epic work type name")
#     parser.add_argument("--task-type", default="Task", help="Jira Task work type name")
#     parser.add_argument("--apply", action="store_true", help="Create/update Jira data")
#     args = parser.parse_args()

#     if not args.workbook.is_file():
#         raise ValueError(f"Workbook not found: {args.workbook}")
#     sheet_name, rows = read_rows(args.workbook, args.sheet)

#     components = unique(row.component for row in rows if row.component)
#     epics = unique(row.epic for row in rows if row.epic)
#     tasks = [row.task for row in rows if row.task]
#     epic_components: dict[str, list[str]] = {
#         epic.casefold(): unique(row.component for row in rows if row.epic.casefold() == epic.casefold())
#         for epic in epics
#     }

#     print(f"Workbook: {args.workbook}")
#     print(f"Sheet: {sheet_name}")
#     print(f"Rows: {len(rows)} | Components: {len(components)} | Epics: {len(epics)} | Tasks: {len(tasks)}")
#     print("\nPlanned hierarchy:")
#     for epic in epics:
#         names = ", ".join(epic_components[epic.casefold()])
#         print(f"  Epic {epic!r} -> Components [{names}]")
#         for row in rows:
#             if row.task and row.epic.casefold() == epic.casefold():
#                 print(f"    Task {row.task!r} -> Parent {epic!r}; Component {row.component!r}")
#     component_only = [row.component for row in rows if row.component and not row.epic and not row.task]
#     for component in component_only:
#         print(f"  Component {component!r} -> no Epics or Tasks (valid)")

#     if not args.apply:
#         print("\nPREVIEW ONLY: Jira was not changed. Add --apply to create/update items.")
#         return 0

#     missing_env = [
#         name for name in ("JIRA_BASE_URL", "JIRA_EMAIL", "JIRA_API_TOKEN") if not os.getenv(name)
#     ]
#     if missing_env:
#         raise ValueError("Missing environment variables: " + ", ".join(missing_env))

#     jira = Jira(os.environ["JIRA_BASE_URL"], os.environ["JIRA_EMAIL"], os.environ["JIRA_API_TOKEN"])
#     project_key = args.project_key.upper()
#     project = jira.project(project_key)
#     project_id = str(project["id"])
#     print(f"\nJira project: {project['key']} - {project['name']} (ID {project_id})")

#     existing_components = jira.components(project_key)
#     component_by_name = {clean(c["name"]).casefold(): c for c in existing_components}
#     for name in components:
#         identity = name.casefold()
#         if identity in component_by_name:
#             print(f"FOUND component: {name}")
#         else:
#             created = jira.create_component(project_key, name)
#             component_by_name[identity] = created
#             print(f"CREATED component: {name}")

#     epic_by_summary = exact_summary_index(jira.issues(project_key, args.epic_type), args.epic_type)
#     for epic in epics:
#         fields = {
#             "project": {"key": project_key},
#             "summary": epic,
#             "issuetype": {"name": args.epic_type},
#             "components": component_ids(epic_components[epic.casefold()], component_by_name),
#         }
#         existing = epic_by_summary.get(epic.casefold())
#         if existing:
#             update_fields = {k: v for k, v in fields.items() if k not in ("project", "issuetype")}
#             jira.update_issue(existing["key"], update_fields)
#             print(f"UPDATED Epic: {existing['key']} - {epic}")
#         else:
#             created = jira.create_issue(fields)
#             existing = {"key": created["key"], "fields": fields}
#             epic_by_summary[epic.casefold()] = existing
#             print(f"CREATED Epic: {created['key']} - {epic}")

#     existing_tasks = jira.issues(project_key, args.task_type)
#     task_candidates: dict[str, list[dict[str, Any]]] = {}
#     for issue in existing_tasks:
#         summary = clean(issue.get("fields", {}).get("summary"))
#         task_candidates.setdefault(summary.casefold(), []).append(issue)
#     processed: set[tuple[str, str, str]] = set()
#     for row in rows:
#         if not row.task:
#             continue
#         epic_issue = epic_by_summary[row.epic.casefold()]
#         identity = (row.task.casefold(), epic_issue["key"].casefold(), row.component.casefold())
#         if identity in processed:
#             continue
#         processed.add(identity)
#         fields = {
#             "project": {"key": project_key},
#             "summary": row.task,
#             "issuetype": {"name": args.task_type},
#             "parent": {"key": epic_issue["key"]},
#             "components": component_ids([row.component], component_by_name),
#         }
#         existing = None
#         desired_component_id = str(component_by_name[row.component.casefold()]["id"])
#         for candidate in task_candidates.get(row.task.casefold(), []):
#             candidate_fields = candidate.get("fields", {})
#             parent_key = clean((candidate_fields.get("parent") or {}).get("key"))
#             candidate_component_ids = {
#                 str(value.get("id")) for value in candidate_fields.get("components", []) if value.get("id")
#             }
#             if parent_key.casefold() == epic_issue["key"].casefold() and desired_component_id in candidate_component_ids:
#                 existing = candidate
#                 break
#         if existing:
#             update_fields = {k: v for k, v in fields.items() if k not in ("project", "issuetype")}
#             jira.update_issue(existing["key"], update_fields)
#             print(
#                 f"UPDATED Task: {existing['key']} - {row.task} "
#                 f"(Parent {epic_issue['key']}, Component {row.component})"
#             )
#         else:
#             created = jira.create_issue(fields)
#             created_issue = {"key": created["key"], "fields": fields}
#             task_candidates.setdefault(row.task.casefold(), []).append(created_issue)
#             print(
#                 f"CREATED Task: {created['key']} - {row.task} "
#                 f"(Parent {epic_issue['key']}, Component {row.component})"
#             )

#     print("\nImport completed successfully.")
#     return 0


# if __name__ == "__main__":
#     try:
#         raise SystemExit(main())
#     except (ValueError, RuntimeError) as exc:
#         print(f"ERROR: {exc}", file=sys.stderr)
#         raise SystemExit(1)


# $env:JIRA_BASE_URL="https://your-site.atlassian.net" 
# $env:JIRA_EMAIL="your-email@example.com"
# $env:JIRA_API_TOKEN="your-api-token"

# python ".\jira_component_epic_task_import.py" "sepm-jira.xlsx" --project-key PER


#!/usr/bin/env python3
"""Import Jira Components, Epics, and Tasks from a three-column Excel sheet.

Expected columns (case/spacing insensitive): components | epics | tasks

The script is a preview unless --apply is supplied. Existing Jira work items are
matched by exact summary inside the selected project and updated, making reruns
safe. Jira credentials are read from environment variables:

    JIRA_BASE_URL, JIRA_EMAIL, JIRA_API_TOKEN
"""

from __future__ import annotations

import argparse
import base64
import json
import os
import re
import sys
import time
from collections import OrderedDict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import load_workbook


def clean(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalized_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", clean(value).casefold())


def unique(values: Iterable[str]) -> list[str]:
    return list(OrderedDict.fromkeys(values))


@dataclass(frozen=True)
class Row:
    excel_row: int
    component: str
    epic: str
    task: str


def prompt_for_value(row_number: int, field: str, context: str) -> str:
    """Ask for a required relationship without changing the workbook."""
    try:
        import tkinter as tk
        from tkinter import simpledialog

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        try:
            while True:
                value = simpledialog.askstring(
                    "Missing Jira mapping",
                    f"Excel row {row_number}: {context}\n\nEnter the missing {field}:",
                    parent=root,
                )
                if value is None:
                    raise ValueError(
                        f"Import cancelled while asking for {field} in Excel row {row_number}."
                    )
                value = clean(value)
                if value:
                    return value
        finally:
            root.destroy()
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError(
            f"Excel row {row_number} requires a {field}, but the input box could not open: {exc}"
        ) from exc


def read_rows(path: Path, sheet_name: str | None) -> tuple[str, list[Row]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"Sheet {sheet_name!r} was not found. Available sheets: "
                    + ", ".join(workbook.sheetnames)
                )
            sheet = workbook[sheet_name]
        else:
            sheet = workbook[workbook.sheetnames[0]]

        # Some valid XLSX files omit the worksheet dimension metadata.
        # In openpyxl read-only mode that makes max_row/max_column None.
        # Force openpyxl to scan the worksheet and determine its real size.
        if sheet.max_row is None or sheet.max_column is None:
            sheet.calculate_dimension(force=True)

        if sheet.max_row is None:
            raise ValueError(f"Could not determine the size of worksheet {sheet.title!r}.")

        header_row = None
        columns: dict[str, int] = {}
        aliases = {
            "components": "component",
            "component": "component",
            "epics": "epic",
            "epic": "epic",
            "tasks": "task",
            "task": "task",
        }
        for row_number, cells in enumerate(
            sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 30), values_only=True),
            start=1,
        ):
            found: dict[str, int] = {}
            for column_number, value in enumerate(cells, start=1):
                canonical = aliases.get(normalized_header(value))
                if canonical:
                    found[canonical] = column_number
            if {"component", "epic", "task"}.issubset(found):
                header_row = row_number
                columns = found
                break

        if header_row is None:
            raise ValueError(
                "Could not find Component, Epic, and Task columns in the first 30 rows."
            )

        rows: list[Row] = []
        seen_task_mappings: set[tuple[str, str, str]] = set()
        for row_number in range(header_row + 1, sheet.max_row + 1):
            component = clean(sheet.cell(row_number, columns["component"]).value)
            epic = clean(sheet.cell(row_number, columns["epic"]).value)
            task = clean(sheet.cell(row_number, columns["task"]).value)
            if not component and not epic and not task:
                continue

            # Component-only rows are valid. Epic rows may omit Task, but every
            # Epic needs a Component. A Task needs both its parent Epic and its
            # Component. Prompt only for relationships that are actually required.
            if epic and not component:
                component = prompt_for_value(
                    row_number, "Component", f"Epic {epic!r} has no Component."
                )
            if task and not epic:
                epic = prompt_for_value(
                    row_number, "Epic", f"Task {task!r} has no parent Epic."
                )
            if task and not component:
                component = prompt_for_value(
                    row_number, "Component", f"Task {task!r} has no Component."
                )

            item = Row(row_number, component, epic, task)
            mapping = (task.casefold(), epic.casefold(), component.casefold())
            if task and mapping in seen_task_mappings:
                continue
            if task:
                seen_task_mappings.add(mapping)
            rows.append(item)

        if not rows:
            raise ValueError("No data rows were found below the headers.")
        return sheet.title, rows
    finally:
        workbook.close()


class Jira:
    def __init__(self, base_url: str, email: str, token: str) -> None:
        self.base_url = base_url.rstrip("/")
        credentials = base64.b64encode(f"{email}:{token}".encode("utf-8")).decode("ascii")
        self.headers = {
            "Authorization": f"Basic {credentials}",
            "Accept": "application/json",
            "Content-Type": "application/json",
        }

    def request(self, method: str, path: str, **kwargs: Any) -> Any:
        params = kwargs.pop("params", None)
        payload = kwargs.pop("json", None)
        if kwargs:
            raise TypeError(f"Unsupported request options: {', '.join(kwargs)}")
        url = self.base_url + path
        if params:
            url += "?" + urlencode(params)
        body = json.dumps(payload).encode("utf-8") if payload is not None else None
        for attempt in range(5):
            try:
                request = Request(url, data=body, headers=self.headers, method=method)
                with urlopen(request, timeout=120) as response:
                    content = response.read()
                    if response.status == 204 or not content:
                        return None
                    return json.loads(content.decode("utf-8"))
            except HTTPError as exc:
                content = exc.read()
                try:
                    detail = json.loads(content.decode("utf-8"))
                except (ValueError, UnicodeDecodeError):
                    detail = content.decode("utf-8", errors="replace")
                if exc.code == 429 or 500 <= exc.code < 600:
                    if attempt < 4:
                        retry_after = exc.headers.get("Retry-After")
                        time.sleep(float(retry_after) if retry_after else 2 ** attempt)
                        continue
                raise RuntimeError(
                    f"Jira {method} {path} returned HTTP {exc.code}: {detail}"
                ) from exc
            except (URLError, TimeoutError) as exc:
                if attempt == 4:
                    raise RuntimeError(f"Jira request failed after retries: {exc}") from exc
                time.sleep(2 ** attempt)
        raise AssertionError("unreachable")

    def project(self, key: str) -> dict[str, Any]:
        return self.request("GET", f"/rest/api/3/project/{key}")

    def components(self, project_key: str) -> list[dict[str, Any]]:
        return self.request("GET", f"/rest/api/3/project/{project_key}/components")

    def create_component(self, project_key: str, name: str) -> dict[str, Any]:
        return self.request(
            "POST", "/rest/api/3/component", json={"name": name, "project": project_key}
        )

    def issues(self, project_key: str, issue_type: str) -> list[dict[str, Any]]:
        # Enhanced search is paginated with nextPageToken.
        jql_type = issue_type.replace('"', '\\"')
        jql = f'project = "{project_key}" AND issuetype = "{jql_type}" ORDER BY key'
        result: list[dict[str, Any]] = []
        token: str | None = None
        while True:
            params: dict[str, Any] = {
                "jql": jql,
                "fields": "summary,components,parent",
                "maxResults": 100,
            }
            if token:
                params["nextPageToken"] = token
            page = self.request("GET", "/rest/api/3/search/jql", params=params)
            result.extend(page.get("issues", []))
            token = page.get("nextPageToken")
            if not token:
                return result

    def create_issue(self, fields: dict[str, Any]) -> dict[str, Any]:
        return self.request("POST", "/rest/api/3/issue", json={"fields": fields})

    def update_issue(self, key: str, fields: dict[str, Any]) -> None:
        self.request("PUT", f"/rest/api/3/issue/{key}", json={"fields": fields})


def exact_summary_index(issues: list[dict[str, Any]], issue_type: str) -> dict[str, dict[str, Any]]:
    index: dict[str, dict[str, Any]] = {}
    for issue in issues:
        summary = clean(issue.get("fields", {}).get("summary"))
        identity = summary.casefold()
        if identity in index:
            raise RuntimeError(
                f"Multiple Jira {issue_type} items have the exact summary {summary!r}. "
                "Rename or remove the duplicate before importing."
            )
        index[identity] = issue
    return index


def component_ids(items: Iterable[str], component_by_name: dict[str, dict[str, Any]]) -> list[dict[str, str]]:
    return [{"id": str(component_by_name[name.casefold()]["id"])} for name in items]


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Import Components, Epics, and Tasks from Excel into Jira Cloud."
    )
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--project-key", required=True)
    parser.add_argument("--sheet", help="Worksheet name; defaults to the first sheet")
    parser.add_argument("--epic-type", default="Epic", help="Jira Epic work type name")
    parser.add_argument("--task-type", default="Task", help="Jira Task work type name")
    parser.add_argument("--apply", action="store_true", help="Create/update Jira data")
    args = parser.parse_args()

    if not args.workbook.is_file():
        raise ValueError(f"Workbook not found: {args.workbook}")
    sheet_name, rows = read_rows(args.workbook, args.sheet)

    components = unique(row.component for row in rows if row.component)
    epics = unique(row.epic for row in rows if row.epic)
    tasks = [row.task for row in rows if row.task]
    epic_components: dict[str, list[str]] = {
        epic.casefold(): unique(row.component for row in rows if row.epic.casefold() == epic.casefold())
        for epic in epics
    }

    print(f"Workbook: {args.workbook}")
    print(f"Sheet: {sheet_name}")
    print(f"Rows: {len(rows)} | Components: {len(components)} | Epics: {len(epics)} | Tasks: {len(tasks)}")
    print("\nPlanned hierarchy:")
    for epic in epics:
        names = ", ".join(epic_components[epic.casefold()])
        print(f"  Epic {epic!r} -> Components [{names}]")
        for row in rows:
            if row.task and row.epic.casefold() == epic.casefold():
                print(f"    Task {row.task!r} -> Parent {epic!r}; Component {row.component!r}")
    component_only = [row.component for row in rows if row.component and not row.epic and not row.task]
    for component in component_only:
        print(f"  Component {component!r} -> no Epics or Tasks (valid)")

    if not args.apply:
        print("\nPREVIEW ONLY: Jira was not changed. Add --apply to create/update items.")
        return 0

    missing_env = [
        name for name in ("JIRA_BASE_URL", "JIRA_EMAIL", "JIRA_API_TOKEN") if not os.getenv(name)
    ]
    if missing_env:
        raise ValueError("Missing environment variables: " + ", ".join(missing_env))

    jira = Jira(os.environ["JIRA_BASE_URL"], os.environ["JIRA_EMAIL"], os.environ["JIRA_API_TOKEN"])
    project_key = args.project_key.upper()
    project = jira.project(project_key)
    project_id = str(project["id"])
    print(f"\nJira project: {project['key']} - {project['name']} (ID {project_id})")

    existing_components = jira.components(project_key)
    component_by_name = {clean(c["name"]).casefold(): c for c in existing_components}
    for name in components:
        identity = name.casefold()
        if identity in component_by_name:
            print(f"FOUND component: {name}")
        else:
            created = jira.create_component(project_key, name)
            component_by_name[identity] = created
            print(f"CREATED component: {name}")

    epic_by_summary = exact_summary_index(jira.issues(project_key, args.epic_type), args.epic_type)
    for epic in epics:
        fields = {
            "project": {"key": project_key},
            "summary": epic,
            "issuetype": {"name": args.epic_type},
            "components": component_ids(epic_components[epic.casefold()], component_by_name),
        }
        existing = epic_by_summary.get(epic.casefold())
        if existing:
            update_fields = {k: v for k, v in fields.items() if k not in ("project", "issuetype")}
            jira.update_issue(existing["key"], update_fields)
            print(f"UPDATED Epic: {existing['key']} - {epic}")
        else:
            created = jira.create_issue(fields)
            existing = {"key": created["key"], "fields": fields}
            epic_by_summary[epic.casefold()] = existing
            print(f"CREATED Epic: {created['key']} - {epic}")

    existing_tasks = jira.issues(project_key, args.task_type)
    task_candidates: dict[str, list[dict[str, Any]]] = {}
    for issue in existing_tasks:
        summary = clean(issue.get("fields", {}).get("summary"))
        task_candidates.setdefault(summary.casefold(), []).append(issue)
    processed: set[tuple[str, str, str]] = set()
    for row in rows:
        if not row.task:
            continue
        epic_issue = epic_by_summary[row.epic.casefold()]
        identity = (row.task.casefold(), epic_issue["key"].casefold(), row.component.casefold())
        if identity in processed:
            continue
        processed.add(identity)
        fields = {
            "project": {"key": project_key},
            "summary": row.task,
            "issuetype": {"name": args.task_type},
            "parent": {"key": epic_issue["key"]},
            "components": component_ids([row.component], component_by_name),
        }
        existing = None
        desired_component_id = str(component_by_name[row.component.casefold()]["id"])
        for candidate in task_candidates.get(row.task.casefold(), []):
            candidate_fields = candidate.get("fields", {})
            parent_key = clean((candidate_fields.get("parent") or {}).get("key"))
            candidate_component_ids = {
                str(value.get("id")) for value in candidate_fields.get("components", []) if value.get("id")
            }
            if parent_key.casefold() == epic_issue["key"].casefold() and desired_component_id in candidate_component_ids:
                existing = candidate
                break
        if existing:
            update_fields = {k: v for k, v in fields.items() if k not in ("project", "issuetype")}
            jira.update_issue(existing["key"], update_fields)
            print(
                f"UPDATED Task: {existing['key']} - {row.task} "
                f"(Parent {epic_issue['key']}, Component {row.component})"
            )
        else:
            created = jira.create_issue(fields)
            created_issue = {"key": created["key"], "fields": fields}
            task_candidates.setdefault(row.task.casefold(), []).append(created_issue)
            print(
                f"CREATED Task: {created['key']} - {row.task} "
                f"(Parent {epic_issue['key']}, Component {row.component})"
            )

    print("\nImport completed successfully.")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (ValueError, RuntimeError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)

